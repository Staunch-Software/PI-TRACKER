import uuid
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import require_roles
from app.core.enums import AuditAction, AuditEntityType, UserRole
from app.db.session import get_db
from app.models.pi_entry import PiEntry
from app.models.user import User
from app.models.vendor import Vendor
from app.models.vessel import Vessel
from app.schemas.import_wizard import ImportCommitRequest, ImportCommitResponse, ImportParseResponse
from app.services.audit import write_audit_log
from app.services.importer import parse_workbook

router = APIRouter(prefix="/import", tags=["import"])


@router.get("/template")
def download_import_template() -> StreamingResponse:
    """Return a pre-formatted .xlsx template that users fill in before uploading."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not available") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PI Follow-up Tracker"

    # ── Headers ──────────────────────────────────────────────────────────────
    headers = [
        "DPR No.",
        "DPR Date",
        "Vessel Name",
        "Vendor Name",
        "Service Details",
        "Amount INR",
        "FC Amount",
        "Currency",
        "Payment Date",
        "Payment Reference",
        "Follow Up Status",
        "Last Known Remark",
        "Reminder 1 Sent Date",
        "Reminder 2 Sent Date",
        "Final Invoice Received",
        "Invoice No",
        "Invoice Date",
        "Notes",
    ]

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    example_fill = PatternFill("solid", fgColor="EBF3FB")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Example row ───────────────────────────────────────────────────────────
    example_row = [
        "DPR-2024-001",      # DPR No.
        "2024-01-15",        # DPR Date
        "MV Example Vessel", # Vessel Name
        "Example Vendor Co.",# Vendor Name
        "Engine Maintenance", # Service Details
        "150000.00",         # Amount INR
        "",                  # FC Amount (optional)
        "INR",               # Currency
        "2024-02-10",        # Payment Date (optional)
        "TXN-REF-001",       # Payment Reference (optional)
        "Pending - Not Yet Followed Up",  # Follow Up Status
        "Awaiting invoice",  # Last Known Remark (optional)
        "",                  # Reminder 1 Sent Date (optional)
        "",                  # Reminder 2 Sent Date (optional)
        "No",                # Final Invoice Received (Yes/No)
        "",                  # Invoice No (optional)
        "",                  # Invoice Date (optional)
        "",                  # Notes (optional)
    ]
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill = example_fill
        cell.font = Font(italic=True, color="555555", name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Data Validation — Follow Up Status ───────────────────────────────────
    status_options = [
        "Pending - Not Yet Followed Up",
        "Pending - Reminder Sent",
        "Pending - Internal Check",
        "Pending - Discrepancy to Resolve",
        "Pending - Scheduled",
        "Pending - Other",
        "Received",
        "Not Applicable",
    ]
    # Write valid statuses to a hidden helper sheet for the dropdown list
    helper_ws = wb.create_sheet("_Lookups")
    helper_ws.sheet_state = "hidden"
    for i, s in enumerate(status_options, start=1):
        helper_ws.cell(row=i, column=1, value=s)
    for i, c in enumerate(["INR", "USD", "EUR"], start=1):
        helper_ws.cell(row=i, column=2, value=c)
    for i, v in enumerate(["Yes", "No"], start=1):
        helper_ws.cell(row=i, column=3, value=v)

    # Follow Up Status column (col 11 = K)
    dv_status = DataValidation(
        type="list",
        formula1=f"'_Lookups'!$A$1:$A${len(status_options)}",
        allow_blank=True,
        showDropDown=False,
    )
    dv_status.error = "Please choose a valid Follow Up Status from the list."
    dv_status.errorTitle = "Invalid Status"
    dv_status.prompt = "Select a Follow Up Status"
    dv_status.promptTitle = "Follow Up Status"
    ws.add_data_validation(dv_status)
    dv_status.sqref = "K3:K10000"

    # Currency column (col 8 = H)
    dv_currency = DataValidation(
        type="list",
        formula1="'_Lookups'!$B$1:$B$3",
        allow_blank=True,
        showDropDown=False,
    )
    dv_currency.error = "Currency must be INR, USD, or EUR."
    dv_currency.errorTitle = "Invalid Currency"
    ws.add_data_validation(dv_currency)
    dv_currency.sqref = "H3:H10000"

    # Final Invoice Received column (col 15 = O)
    dv_bool = DataValidation(
        type="list",
        formula1="'_Lookups'!$C$1:$C$2",
        allow_blank=True,
        showDropDown=False,
    )
    dv_bool.error = "Please enter Yes or No."
    dv_bool.errorTitle = "Invalid Value"
    ws.add_data_validation(dv_bool)
    dv_bool.sqref = "O3:O10000"

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [16, 14, 22, 22, 28, 14, 12, 10, 14, 22, 32, 28, 22, 22, 22, 16, 14, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20

    # Freeze the header row
    ws.freeze_panes = "A3"

    # ── Instructions sheet ────────────────────────────────────────────────────
    info_ws = wb.create_sheet("Instructions", 0)
    info_ws.sheet_view.showGridLines = False
    instructions = [
        ("PI Follow-up Tracker — Import Template", True, "1E3A5F", 16),
        ("", False, "000000", 11),
        ("HOW TO USE THIS TEMPLATE", True, "1E3A5F", 12),
        ("1. Go to the 'PI Follow-up Tracker' sheet.", False, "222222", 11),
        ("2. Fill in your data starting from row 3 (row 2 is just an example — you can delete it).", False, "222222", 11),
        ("3. Required columns: DPR No., DPR Date, Vessel Name, Vendor Name.", False, "CC0000", 11),
        ("4. All other columns are optional.", False, "222222", 11),
        ("5. Use the dropdown lists in the Currency, Follow Up Status, and Final Invoice Received columns.", False, "222222", 11),
        ("6. Dates can be entered as YYYY-MM-DD, DD/MM/YYYY, or any standard Excel date format.", False, "222222", 11),
        ("7. Save the file as .xlsx and upload it using the Import button in the application.", False, "222222", 11),
        ("", False, "000000", 11),
        ("COLUMN REFERENCE", True, "1E3A5F", 12),
        ("DPR No.          — Unique identifier for each PI entry (e.g. DPR-2024-001). REQUIRED.", False, "222222", 10),
        ("DPR Date         — Date the DPR was raised. REQUIRED.", False, "222222", 10),
        ("Vessel Name      — Name of the vessel. REQUIRED.", False, "222222", 10),
        ("Vendor Name      — Name of the vendor / supplier. REQUIRED.", False, "222222", 10),
        ("Service Details  — Short description of the service provided.", False, "555555", 10),
        ("Amount INR       — Invoice amount in INR (numeric, e.g. 150000.00).", False, "555555", 10),
        ("FC Amount        — Foreign currency amount if applicable.", False, "555555", 10),
        ("Currency         — INR / USD / EUR. Defaults to INR if blank.", False, "555555", 10),
        ("Payment Date     — Date payment was made.", False, "555555", 10),
        ("Payment Reference— Bank/transaction reference number.", False, "555555", 10),
        ("Follow Up Status — Select from dropdown list.", False, "555555", 10),
        ("Last Known Remark— Free text remark.", False, "555555", 10),
        ("Reminder 1/2     — Dates reminder emails were sent.", False, "555555", 10),
        ("Final Invoice Received — Yes or No.", False, "555555", 10),
        ("Invoice No.      — Invoice number once received.", False, "555555", 10),
        ("Invoice Date     — Date on the invoice.", False, "555555", 10),
        ("Notes            — Any additional notes.", False, "555555", 10),
    ]
    for row_idx, (text, bold, color, size) in enumerate(instructions, start=1):
        cell = info_ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, color=color, name="Calibri", size=size)
        cell.alignment = Alignment(wrap_text=True)
    info_ws.column_dimensions["A"].width = 90

    # Make the Instructions sheet the active one when the file opens
    wb.active = info_ws

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="PI_Import_Template.xlsx"'},
    )


@router.post("/parse", response_model=ImportParseResponse)
async def parse_import_file(
    file: UploadFile,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN, UserRole.EDITOR)),
) -> dict:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Please upload an .xlsx file")

    contents = await file.read()
    try:
        rows = parse_workbook(db, contents)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Could not read file: {exc}") from exc

    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No data rows found in the file")

    return {
        "rows": rows,
        "total_rows": len(rows),
        "valid_rows": sum(1 for r in rows if not r.errors),
        "error_rows": sum(1 for r in rows if r.errors),
        "duplicate_rows": sum(1 for r in rows if r.is_duplicate),
    }


def _get_or_create_vessel(db: Session, name: str, created_by) -> Vessel:
    vessel = db.query(Vessel).filter(Vessel.name.ilike(name)).first()
    if not vessel:
        vessel = Vessel(name=name, created_by=created_by)
        db.add(vessel)
        db.flush()
    return vessel


def _get_or_create_vendor(db: Session, name: str, created_by) -> Vendor:
    vendor = db.query(Vendor).filter(Vendor.name.ilike(name)).first()
    if not vendor:
        vendor = Vendor(name=name, created_by=created_by)
        db.add(vendor)
        db.flush()
    return vendor


@router.post("/commit", response_model=ImportCommitResponse)
def commit_import(
    payload: ImportCommitRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN, UserRole.EDITOR)),
) -> dict:
    inserted = updated = skipped = 0
    errors: list[str] = []

    entry_fields = [
        "dpr_no",
        "dpr_date",
        "service_details",
        "amount_inr",
        "fc_amount",
        "currency",
        "payment_date",
        "payment_reference",
        "followup_status",
        "last_known_remark",
        "reminder_1_sent_date",
        "reminder_2_sent_date",
        "final_invoice_received",
        "invoice_no",
        "invoice_date",
        "notes",
    ]

    for row in payload.rows:
        if row.decision == "skip":
            skipped += 1
            continue

        try:
            with db.begin_nested():
                vessel = _get_or_create_vessel(db, row.vessel_name, current_user.id)
                vendor = _get_or_create_vendor(db, row.vendor_name, current_user.id)
                values = {f: getattr(row, f) for f in entry_fields}
                values["vessel_id"] = vessel.id
                values["vendor_id"] = vendor.id

                if row.decision == "insert":
                    if db.query(PiEntry).filter(PiEntry.dpr_no == row.dpr_no).first():
                        raise ValueError(f"DPR No. '{row.dpr_no}' already exists (row {row.row_number})")
                    db.add(PiEntry(**values, created_by=current_user.id))
                    inserted += 1
                else:  # update
                    entry = db.get(PiEntry, row.existing_id) if row.existing_id else None
                    if not entry:
                        raise ValueError(f"Row {row.row_number}: existing entry not found for update")
                    for f, v in values.items():
                        setattr(entry, f, v)
                    entry.updated_by = current_user.id
                    updated += 1
        except Exception as exc:
            errors.append(str(exc))

    if inserted or updated:
        write_audit_log(
            db,
            entity_type=AuditEntityType.IMPORT_BATCH,
            entity_id=uuid.uuid4(),
            action=AuditAction.IMPORT,
            changed_by=current_user.id,
            summary=(
                f"{current_user.full_name} imported {inserted + updated} PI entries from Excel "
                f"({inserted} added, {updated} updated, {skipped} skipped)"
            ),
        )
    db.commit()

    return {"inserted": inserted, "updated": updated, "skipped": skipped, "failed": len(errors), "errors": errors}
