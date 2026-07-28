import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from dateutil import parser as dateutil_parser
from sqlalchemy.orm import Session

from app.core.enums import FOLLOW_UP_STATUS_LABELS, Currency, FollowUpStatus
from app.models.pi_entry import PiEntry
from app.models.vendor import Vendor
from app.models.vessel import Vessel
from app.schemas.import_wizard import ImportRowPreview

# Canonical 23-sheet-column headers -> our field names, keyed by normalized header text
# (lowercased, punctuation stripped to spaces, whitespace collapsed). Both the "Vessel"/"Vessel
# Name" style short and long forms are accepted since real-world sheets vary. "S.No." and "Days
# Since Payment" are recognized but always ignored (derived fields); invoice_file_name/
# attached_by/date_attached are attachment-provenance fields only meaningful once a real file is
# attached (Phase 3), so they are not imported here even though the sheet has columns for them.
_HEADER_MAP = {
    "dpr no": "dpr_no",
    "dpr date": "dpr_date",
    "vessel": "vessel_name",
    "vessel name": "vessel_name",
    "vendor": "vendor_name",
    "vendor name": "vendor_name",
    "service details": "service_details",
    "amount inr": "amount_inr",
    "fc amount": "fc_amount",
    "currency": "currency",
    "payment date": "payment_date",
    "payment reference": "payment_reference",
    "follow up status": "followup_status",
    "last known remark": "last_known_remark",
    "reminder 1 sent date": "reminder_1_sent_date",
    "reminder 1 sent": "reminder_1_sent_date",
    "reminder 2 sent date": "reminder_2_sent_date",
    "reminder 2 sent": "reminder_2_sent_date",
    "final invoice received": "final_invoice_received",
    "final invoice received y n": "final_invoice_received",
    "invoice no": "invoice_no",
    "invoice date": "invoice_date",
    "notes": "notes",
}

_IGNORED_HEADERS = {"s no", "days since payment", "invoice file name", "attached by", "date attached"}

# Distinctive header text used to locate the real header row — source sheets often have title/
# subtitle rows above it (e.g. "Proforma Invoice - Final Tax Invoice Follow-up Tracker").
_HEADER_ROW_SIGNAL = "dpr no"
_MAX_HEADER_SEARCH_ROWS = 15


def _normalize_header(raw: str) -> str:
    lowered = re.sub(r"[^a-z0-9\s]", " ", raw.lower())
    return re.sub(r"\s+", " ", lowered).strip()


def _parse_date(value) -> tuple[date | None, str | None]:
    if value is None or value == "":
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    try:
        return dateutil_parser.parse(str(value), dayfirst=True).date(), None
    except (ValueError, OverflowError):
        return None, f"Unparseable date: '{value}'"


def _parse_amount(value) -> tuple[Decimal | None, str | None]:
    if value is None or value == "":
        return None, None
    try:
        return Decimal(str(value)), None
    except InvalidOperation:
        return None, f"Not a number: '{value}'"


def _parse_currency(value) -> tuple[Currency, str | None]:
    if value is None or str(value).strip() == "":
        return Currency.INR, None
    normalized = str(value).strip().upper()
    if normalized in Currency.__members__:
        return Currency[normalized], None
    return Currency.INR, f"Unknown currency '{value}' (expected INR/USD/EUR)"


def _parse_status(value) -> tuple[FollowUpStatus, str | None]:
    if value is None or str(value).strip() == "":
        return FollowUpStatus.PENDING_NOT_YET_FOLLOWED_UP, None
    normalized = str(value).strip().lower()
    for status, label in FOLLOW_UP_STATUS_LABELS.items():
        if label.lower() == normalized:
            return status, None
    return FollowUpStatus.PENDING_NOT_YET_FOLLOWED_UP, f"Unknown follow-up status '{value}'"


def _parse_bool(value) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in ("true", "yes", "y", "1")


def _find_header_row(sheet) -> int | None:
    """Scans the first few rows for the one containing the DPR No. column — real-world sheets
    often have title/subtitle rows above the actual header (see PI_Followup_Tracker.xlsx)."""
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=_MAX_HEADER_SEARCH_ROWS, values_only=True), start=1):
        normalized_cells = {_normalize_header(str(cell)) for cell in row if cell}
        if _HEADER_ROW_SIGNAL in normalized_cells:
            return row_idx
    return None


def parse_workbook(db: Session, file_bytes: bytes) -> list[ImportRowPreview]:
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    sheet = None
    header_row_idx = None
    for candidate in workbook.worksheets:
        found_at = _find_header_row(candidate)
        if found_at:
            sheet, header_row_idx = candidate, found_at
            break

    if sheet is None or header_row_idx is None:
        raise ValueError("Could not find a 'DPR No.' header row in any sheet of this workbook")

    header_row = next(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    field_by_col: dict[int, str] = {}
    for col_idx, raw_header in enumerate(header_row):
        if not raw_header:
            continue
        normalized = _normalize_header(str(raw_header))
        if normalized in _IGNORED_HEADERS:
            continue
        field = _HEADER_MAP.get(normalized)
        if field:
            field_by_col[col_idx] = field

    existing_dpr_nos = {row[0]: row[1] for row in db.query(PiEntry.dpr_no, PiEntry.id).all()}
    existing_vessel_names = {v.lower() for (v,) in db.query(Vessel.name).all()}
    existing_vendor_names = {v.lower() for (v,) in db.query(Vendor.name).all()}

    previews: list[ImportRowPreview] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        raw: dict[str, object] = {}
        for col_idx, field in field_by_col.items():
            raw[field] = row[col_idx] if col_idx < len(row) else None

        errors: list[str] = []

        dpr_no = str(raw.get("dpr_no") or "").strip() or None
        if not dpr_no:
            errors.append("DPR No. is required")

        dpr_date, err = _parse_date(raw.get("dpr_date"))
        if err:
            errors.append(f"DPR Date: {err}")
        payment_date, err = _parse_date(raw.get("payment_date"))
        if err:
            errors.append(f"Payment Date: {err}")
        reminder_1, err = _parse_date(raw.get("reminder_1_sent_date"))
        if err:
            errors.append(f"Reminder 1 Sent Date: {err}")
        reminder_2, err = _parse_date(raw.get("reminder_2_sent_date"))
        if err:
            errors.append(f"Reminder 2 Sent Date: {err}")
        invoice_date, err = _parse_date(raw.get("invoice_date"))
        if err:
            errors.append(f"Invoice Date: {err}")

        amount_inr, err = _parse_amount(raw.get("amount_inr"))
        if err:
            errors.append(f"Amount (INR): {err}")
        fc_amount, err = _parse_amount(raw.get("fc_amount"))
        if err:
            errors.append(f"FC Amount: {err}")

        currency, err = _parse_currency(raw.get("currency"))
        if err:
            errors.append(err)
        followup_status, err = _parse_status(raw.get("followup_status"))
        if err:
            errors.append(err)

        vessel_name = str(raw.get("vessel_name") or "").strip() or None
        vendor_name = str(raw.get("vendor_name") or "").strip() or None
        if not vessel_name:
            errors.append("Vessel is required")
        if not vendor_name:
            errors.append("Vendor is required")

        previews.append(
            ImportRowPreview(
                row_number=row_idx,
                dpr_no=dpr_no,
                dpr_date=dpr_date,
                vessel_name=vessel_name,
                vendor_name=vendor_name,
                service_details=(str(raw["service_details"]).strip() if raw.get("service_details") else None),
                amount_inr=amount_inr,
                fc_amount=fc_amount,
                currency=currency,
                payment_date=payment_date,
                payment_reference=(str(raw["payment_reference"]).strip() if raw.get("payment_reference") else None),
                followup_status=followup_status,
                last_known_remark=(str(raw["last_known_remark"]).strip() if raw.get("last_known_remark") else None),
                reminder_1_sent_date=reminder_1,
                reminder_2_sent_date=reminder_2,
                final_invoice_received=_parse_bool(raw.get("final_invoice_received")),
                invoice_no=(str(raw["invoice_no"]).strip() if raw.get("invoice_no") else None),
                invoice_date=invoice_date,
                notes=(str(raw["notes"]).strip() if raw.get("notes") else None),
                errors=errors,
                is_duplicate=dpr_no in existing_dpr_nos if dpr_no else False,
                existing_id=existing_dpr_nos.get(dpr_no) if dpr_no else None,
                vessel_exists=(vessel_name or "").lower() in existing_vessel_names,
                vendor_exists=(vendor_name or "").lower() in existing_vendor_names,
            )
        )

    return previews
